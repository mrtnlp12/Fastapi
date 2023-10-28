from os import path, getcwd
from openpyxl import load_workbook
from schemas.metadata import metadataEntity
from models.metadata import Metadata
from pymongo import MongoClient
from services import email


def update_report(metadata: Metadata, cnn: MongoClient):
    report_path = path.join(
        "/Users",
        "aldocitalan",
        "Desktop",
        "ALZ",
        "Concentrado General",
        "concentrado.xlsx",
    )
    reportWb = load_workbook(report_path)
    datasource_ws = load_workbook(path.join(getcwd(), "fuente.xlsx")).active

    report_ws = reportWb.active

    metadata = metadataEntity(cnn.local.metadata.find_one())

    # modificar un valor de la metadata

    fuente_row = metadata["fuenteTop"]
    report_row = metadata["concentradoTop"]

    column = 1
    inserted = 0

    while True:
        FECHA = datasource_ws.cell(row=fuente_row, column=column).value
        if FECHA is None:
            metadata["fuenteTop"] = fuente_row
            metadata["concentradoTop"] = report_row
            break

        column += 1
        REMISION = datasource_ws.cell(row=fuente_row, column=column).value
        column += 1
        M3 = datasource_ws.cell(row=fuente_row, column=column).value
        column += 1
        NUMERO_DE_UNIDAD = datasource_ws.cell(row=fuente_row, column=column).value
        column += 1
        OPERADOR = datasource_ws.cell(row=fuente_row, column=column).value
        column += 1
        CLIENTE = datasource_ws.cell(row=fuente_row, column=column).value
        column += 1
        DIRECCION = datasource_ws.cell(row=fuente_row, column=column).value
        column += 1
        TIPO_DE_CONCRETO = datasource_ws.cell(row=fuente_row, column=column).value
        column += 1
        CEMENTO = datasource_ws.cell(row=fuente_row, column=column).value
        column += 2
        GRAVAS = datasource_ws.cell(row=fuente_row, column=column).value
        column += 3
        ARENAS = datasource_ws.cell(row=fuente_row, column=column).value
        column += 1
        AGUA = datasource_ws.cell(row=fuente_row, column=column).value
        column += 1
        ADITIVOS = (
            datasource_ws.cell(row=fuente_row, column=column).value,
            datasource_ws.cell(row=fuente_row + 1, column=column).value,
        )
        column += 3
        TIRO = datasource_ws.cell(row=fuente_row, column=column).value
        column += 1
        HORA_SALIDA_PLANTA = datasource_ws.cell(row=fuente_row, column=column).value
        column += 1
        VENDEDOR = datasource_ws.cell(row=fuente_row, column=column).value

        column = 1

        report_ws.cell(row=report_row, column=column).value = FECHA
        column += 1
        report_ws.cell(row=report_row, column=column).value = REMISION
        column += 1
        report_ws.cell(row=report_row, column=column).value = "M3"
        column += 1
        report_ws.cell(row=report_row, column=column).value = M3
        column += 3
        report_ws.cell(row=report_row, column=column).value = DIRECCION
        column += 1
        report_ws.cell(row=report_row, column=column).value = TIPO_DE_CONCRETO
        column += 1
        report_ws.cell(row=report_row, column=column).value = CLIENTE
        column += 6
        report_ws.cell(row=report_row, column=column).value = VENDEDOR
        column += 1
        report_ws.cell(row=report_row, column=column).value = NUMERO_DE_UNIDAD
        column += 1
        report_ws.cell(row=report_row, column=column).value = OPERADOR

        column = 1
        fuente_row += 1
        report_row += 1
        inserted += 1

    metadata["insertados"] = inserted

    cnn.local.metadata.update_one({}, {"$set": metadata})

    reportWb.save(report_path)
    email.EmailService().filesUpdated.append(report_path)
    return


def generate_daily_report(metadata, cnn: MongoClient):
    dailyreport_path = path.join(
        "/Users",
        "aldocitalan",
        "Desktop",
        "ALZ",
        "Plantillas",
        "reporteDiario.xlsx",
    )

    datasource_ws = load_workbook(path.join(getcwd(), "fuente.xlsx")).active
    daily_report_wb = load_workbook(dailyreport_path)
    daily_report_ws = daily_report_wb.active

    fuente_row = metadata["fuenteTop"]
    report_row = 5

    filename = str(datasource_ws.cell(row=fuente_row, column=1).value)

    if path.exists(
        path.join(
            "/Users",
            "aldocitalan",
            "Desktop",
            "ALZ",
            "Reporte Diario",
            f"{filename}.xlsx",
        )
    ):
        return
    while True:
        fecha = datasource_ws.cell(row=fuente_row, column=1).value
        if fecha is None:
            break
        folio = datasource_ws.cell(row=fuente_row, column=2).value
        volumen = datasource_ws.cell(row=fuente_row, column=3).value
        numero_unidad = datasource_ws.cell(row=fuente_row, column=4).value
        cliente = datasource_ws.cell(row=fuente_row, column=6).value
        direccion = datasource_ws.cell(row=fuente_row, column=7).value
        tipo_concreto = datasource_ws.cell(row=fuente_row, column=8).value
        vendedor = datasource_ws.cell(row=fuente_row, column=21).value

        daily_report_ws.cell(row=report_row, column=2).value = fecha
        daily_report_ws.cell(row=report_row, column=3).value = cliente
        daily_report_ws.cell(row=report_row, column=4).value = direccion
        daily_report_ws.cell(row=report_row, column=5).value = "ALZ"
        daily_report_ws.cell(row=report_row, column=6).value = folio
        daily_report_ws.cell(row=report_row, column=7).value = volumen
        daily_report_ws.cell(row=report_row, column=8).value = "M3"
        daily_report_ws.cell(row=report_row, column=14).value = tipo_concreto
        daily_report_ws.cell(row=report_row, column=22).value = vendedor
        daily_report_ws.cell(row=report_row, column=23).value = numero_unidad
        daily_report_ws.cell(row=report_row, column=24).value = numero_unidad

        fuente_row += 1
        report_row += 1

    if report_row > 5:
        daily_report_wb.save(
            path.join(
                "/Users",
                "aldocitalan",
                "Desktop",
                "ALZ",
                "Reporte Diario",
                f"{filename}.xlsx",
            )
        )
        email.EmailService().filesCreated.append(
            path.join(
                "/Users",
                "aldocitalan",
                "Desktop",
                "ALZ",
                "Reporte Diario",
                f"{filename}.xlsx",
            )
        )
    return


def generate_report_by_client(metadata: Metadata, cnn: MongoClient):
    fuente_row = metadata["fuenteTop"]
    datasource_ws = load_workbook(path.join(getcwd(), "fuente.xlsx")).active
    while True:
        fecha = datasource_ws.cell(row=fuente_row, column=1).value
        if fecha == None:
            break
        folio = datasource_ws.cell(row=fuente_row, column=2).value
        unidad_medida = "M3"
        volumen = datasource_ws.cell(row=fuente_row, column=3).value
        direccion = datasource_ws.cell(row=fuente_row, column=7).value
        material = datasource_ws.cell(row=fuente_row, column=8).value
        cliente = datasource_ws.cell(row=fuente_row, column=6).value

        client_path = path.join(
            "/Users",
            "aldocitalan",
            "Desktop",
            "ALZ",
            "Clientes",
            f"{fecha.year}",
            f"{cliente.split(': ')[-1]}.xlsx",
        )

        default_template = path.join(
            "/Users", "aldocitalan", "Desktop", "ALZ", "Plantillas", "cliente.xlsx"
        )

        if path.exists(client_path):
            email.EmailService().filesUpdated.append(client_path)
            client_wb = load_workbook(client_path)
        else:
            email.EmailService().filesCreated.append(client_path)
            client_wb = load_workbook(default_template)

        if cliente not in [cliente["nombre"] for cliente in metadata["clientes"]]:
            metadata["clientes"].append({"nombre": cliente, "fila": 4})

        client_row_and_index = {"row": 0, "index": 0}
        for index, dict_cliente in enumerate(metadata["clientes"]):
            if dict_cliente["nombre"] == cliente:
                client_row_and_index["row"] = dict_cliente["fila"]
                client_row_and_index["index"] = index
                break

        client_wb.active.cell(row=client_row_and_index["row"], column=1).value = fecha
        client_wb.active.cell(row=client_row_and_index["row"], column=2).value = folio
        client_wb.active.cell(
            row=client_row_and_index["row"], column=3
        ).value = unidad_medida
        client_wb.active.cell(row=client_row_and_index["row"], column=4).value = volumen
        client_wb.active.cell(
            row=client_row_and_index["row"], column=7
        ).value = direccion
        client_wb.active.cell(
            row=client_row_and_index["row"], column=8
        ).value = material
        client_wb.active.cell(row=client_row_and_index["row"], column=9).value = cliente

        metadata["clientes"][client_row_and_index["index"]]["fila"] += 1

        fuente_row += 1

        client_wb.save(client_path)

    cnn.local.metadata.update_one({}, {"$set": metadata})

    return
